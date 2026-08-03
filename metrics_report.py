"""Metrics report pipeline: pilot Postgres -> one Excel workbook with dashboards.

Run from the bundle folder with the venv active:

    python metrics_report.py                     # -> reports/KnowledgeHub_Metrics_<date>.xlsx
    python metrics_report.py --out my_report.xlsx

What it does, every run (re-runnable; each run is a fresh snapshot):
  1. Queries the pilot DB's observability tables — extraction_runs,
     pending_facts, quarantined_extractions, resolution_decisions,
     match_candidates, labels, review_queue.
  2. Dumps them into Raw_* sheets (the snapshot layer — plain data, no formulas).
  3. Builds a Metrics sheet of SUMIFS/COUNTIFS formulas over the Raw_* sheets
     (edit or extend a Raw_ sheet and every metric recomputes in Excel).
  4. Builds a Dashboard sheet: KPI tiles + native charts wired to Metrics cells.

Forward-compatibility: when migration 006 lands (benchmark_runs / gold_sets,
per the benchmark methodology doc), the script detects the tables and adds a
Raw_BenchmarkRuns sheet automatically; until then the README notes their absence.

Conventions: DSN comes from knowledge_hub settings (same as check_stack.py);
extractor_version values without a '/pN' suffix predate the contract-suffix
convention and are labeled contract 'p1'.
"""
from __future__ import annotations

import argparse
import datetime as dt
import subprocess
import sys
import tempfile
from pathlib import Path

import psycopg
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_VERSION = "0.1.0"

# Chart palette (dataviz reference palette, light mode, categorical slots 1-3;
# chrome colors from the same reference). Slot order is the CVD-safety
# mechanism — assign in order, never cycle.
SERIES_1 = "2A78D6"   # blue
SERIES_2 = "EB6834"   # orange
SERIES_3 = "1BAF7A"   # aqua
INK      = "0B0B0B"
INK_2    = "52514E"
MUTED    = "898781"
HEADER_FILL = "F0EFEC"   # table headers / section banners
TILE_FILL   = "F9F9F7"   # KPI tile surface (palette page plane)
GRID        = "E1E0D9"   # hairline borders

FONT = "Arial"

HAIRLINE = Side(style="thin", color=GRID)
ACCENT_TOP = Side(style="medium", color=SERIES_1)

# ---------------------------------------------------------------------------
# Queries — one per Raw_ sheet. Column order here IS the sheet column order;
# the Metrics formulas below reference these columns by letter, so the two
# must move together.
# ---------------------------------------------------------------------------
CONTRACT_SQL = ("CASE WHEN position('/' in extractor_version) > 0 "
                "THEN split_part(extractor_version, '/', 2) ELSE 'p1' END")

RAW_SHEETS: dict[str, tuple[list[str], str]] = {
    "Raw_ExtractionRuns": (
        ["id", "created_at", "document_id", "source_chunk_id", "strategy",
         "extractor", "extractor_version", "contract", "ontology_version",
         "prompt_tokens", "output_tokens", "wall_ms", "facts_staged",
         "mentions_staged", "quarantined", "grounding_flags", "repairs",
         "status"],
        f"""SELECT id, created_at, document_id, source_chunk_id, strategy,
                   extractor, extractor_version, {CONTRACT_SQL},
                   ontology_version, prompt_tokens, output_tokens, wall_ms,
                   facts_staged, mentions_staged, quarantined,
                   grounding_flags, repairs, status
            FROM extraction_runs WHERE tenant_id = %(tenant)s ORDER BY id""",
    ),
    "Raw_PendingFacts": (
        ["id", "subject_ref", "predicate", "object", "grounding",
         "needs_review", "confidence", "serialized_lines", "oversized",
         "extractor_version", "contract", "source_chunk_id",
         "resolution_status"],
        f"""SELECT id, subject_ref, predicate,
                   coalesce(object_ref, object_literal), grounding,
                   needs_review, confidence, serialized_lines, oversized,
                   extractor_version, {CONTRACT_SQL}, source_chunk_id,
                   resolution_status
            FROM pending_facts WHERE tenant_id = %(tenant)s ORDER BY id""",
    ),
    "Raw_Quarantine": (
        ["id", "created_at", "reason", "detail", "extractor",
         "extractor_version", "contract", "status", "source_chunk_id"],
        f"""SELECT id, created_at, reason, detail, extractor,
                   extractor_version, {CONTRACT_SQL}, status, source_chunk_id
            FROM quarantined_extractions WHERE tenant_id = %(tenant)s ORDER BY id""",
    ),
    "Raw_Resolution": (
        ["id", "created_at", "mention_id", "tier", "method", "score", "band",
         "decision", "entity_id", "resolver_version", "wall_ms"],
        """SELECT id, created_at, mention_id, tier, method, score, band,
                  decision, entity_id, resolver_version, wall_ms
           FROM resolution_decisions WHERE tenant_id = %(tenant)s ORDER BY id""",
    ),
    "Raw_MatchCandidates": (
        ["id", "created_at", "left_type", "left_id", "right_type", "right_id",
         "match_score", "match_method", "band", "decision"],
        """SELECT id, created_at, left_type, left_id, right_type, right_id,
                  match_score, match_method, band, decision
           FROM match_candidates WHERE tenant_id = %(tenant)s ORDER BY id""",
    ),
    "Raw_Labels": (
        ["id", "created_at", "label_type", "source", "authority",
         "confidence", "ontology_version"],
        """SELECT id, created_at, label_type, source, authority, confidence,
                  ontology_version
           FROM labels WHERE tenant_id = %(tenant)s ORDER BY id""",
    ),
    "Raw_ReviewQueue": (
        ["kind", "ref_id", "context", "created_at"],
        """SELECT kind, ref_id, context, created_at
           FROM review_queue WHERE tenant_id = %(tenant)s ORDER BY created_at, kind, ref_id""",
    ),
}

# Present only once migration 006 (benchmark harness) is applied.
BENCHMARK_SHEETS: dict[str, tuple[list[str], str]] = {
    "Raw_BenchmarkRuns": (
        ["id", "started_at", "finished_at", "tenant_id", "axis",
         "config_label", "gold_set", "pin_profile", "advisory", "status",
         "wall_ms", "headline_name", "headline_value", "package_version",
         "runner_version", "code_hash"],
        """SELECT r.id, r.started_at, r.finished_at, r.tenant_id, r.axis,
                  coalesce(r.config->>'label', r.config->>'embedder',
                           r.config->>'index', r.config->>'scorer',
                           r.config->>'model'),
                  g.kind || '/' || g.version, r.pin_profile_name, r.advisory,
                  r.status, r.wall_ms, r.metrics->>'headline_name',
                  (r.metrics->>'headline_value')::real, r.package_version,
                  r.runner_version, left(r.code_hash, 12)
           FROM benchmark_runs r JOIN gold_sets g ON g.id = r.gold_set_id
           ORDER BY r.id""",
    ),
    # Comparability keys (gold_set, pin_profile) ride every row; the
    # Benchmarks sheet charts one (axis, gold_set, pin_profile) group each —
    # never across groups.
    "Raw_Leaderboard": (
        ["tenant_id", "axis", "gold_set", "pin_profile", "run_id",
         "config", "headline_name", "headline_value", "advisory",
         "superseded", "wall_ms", "finished_at"],
        """SELECT tenant_id, axis, gold_set_kind || '/' || gold_set_version,
                  pin_profile_name, run_id,
                  config_label || CASE WHEN advisory THEN ' (advisory)'
                                       ELSE '' END
                               || CASE WHEN superseded THEN ' (superseded)'
                                       ELSE '' END,
                  headline_name, headline_value, advisory, superseded,
                  wall_ms, finished_at
           FROM benchmark_leaderboard
           ORDER BY axis, gold_set_id, pin_profile_name,
                    headline_value DESC NULLS LAST""",
    ),
}


def _dsn() -> str:
    try:
        from knowledge_hub import settings
        return settings.postgres_dsn
    except Exception:
        return "host=localhost port=5432 dbname=knowledge_hub user=kh password=kh_pilot_pw"


def _clean(v):
    """Make a DB value openpyxl-safe (no tz-aware datetimes, no Decimals)."""
    import decimal
    if isinstance(v, dt.datetime):
        return v.astimezone().replace(tzinfo=None, microsecond=0)
    if isinstance(v, decimal.Decimal):
        return float(v)
    return v


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------
def style_header_row(ws, ncols: int, row: int = 1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=10, bold=True, color=INK)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = Border(bottom=HAIRLINE)


def sheet_title(ws, title: str, subtitle: str = "", span: int = 13):
    """Consistent page header: 16pt title, muted subtitle, hairline rule."""
    put(ws, 1, 1, title, bold=True, size=16)
    ws.row_dimensions[1].height = 26
    if subtitle:
        put(ws, 2, 1, subtitle, color=INK_2, size=9)
    for c in range(1, span + 1):
        ws.cell(row=2, column=c).border = Border(bottom=HAIRLINE)


def draw_tile(ws, row: int, col: int, label: str, formula: str, fmt: str):
    """A KPI tile: 2x2 merged block, page-plane fill, hairline ring, blue
    accent on top. Label in muted ink above a large value."""
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col,
                   end_row=row + 1, end_column=col + 1)
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = Font(name=FONT, size=9, color=INK_2)
    label_cell.alignment = Alignment(horizontal="left", vertical="bottom",
                                     indent=1)
    value_cell = ws.cell(row=row + 1, column=col, value=formula)
    value_cell.font = Font(name=FONT, size=22, bold=True, color=INK)
    value_cell.number_format = fmt
    value_cell.alignment = Alignment(horizontal="left", vertical="top",
                                     indent=1)
    fill = PatternFill("solid", fgColor=TILE_FILL)
    for r, c in ((row, col), (row, col + 1), (row + 1, col), (row + 1, col + 1)):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill
        cell.border = Border(
            top=ACCENT_TOP if r == row else None,
            bottom=HAIRLINE if r == row + 1 else None,
            left=HAIRLINE if c == col else None,
            right=HAIRLINE if c == col + 1 else None)


def write_raw_sheet(wb: Workbook, name: str, headers: list[str], rows) -> int:
    ws = wb.create_sheet(name)
    ws.append(headers)
    style_header_row(ws, len(headers))
    n = 0
    for r in rows:
        ws.append([_clean(v) for v in r])
        n += 1
    datetime_cols: set[int] = set()
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=FONT, size=10)
            if isinstance(cell.value, dt.datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
                datetime_cols.add(cell.column)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = (
            17 if i in datetime_cols else max(12, min(len(h) + 4, 24)))
    ws.freeze_panes = "A2"
    return n


def section_title(ws, row: int, text: str, span: int = 5):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=12, bold=True, color=INK)
    ws.row_dimensions[row].height = 22
    c.alignment = Alignment(vertical="bottom")
    # Blue tick + hairline rule under the title, spanning the section width.
    c.border = Border(bottom=Side(style="thick", color=SERIES_1))
    for col in range(2, span + 1):
        ws.cell(row=row, column=col).border = Border(bottom=HAIRLINE)


def put(ws, row: int, col: int, value, *, bold=False, fmt=None, color=INK, size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    if fmt:
        cell.number_format = fmt
    return cell


def build_metrics_sheet(wb: Workbook):
    """Everything here is a FORMULA over the Raw_ sheets — no computed values."""
    ws = wb.create_sheet("Metrics")
    ws.column_dimensions["A"].width = 34
    for col in "BCDE":
        ws.column_dimensions[col].width = 14

    ER = "Raw_ExtractionRuns"
    N = "#,##0"
    PCT = "0.0%"
    DEC = "#,##0.0"

    # --- Section A: extraction contract comparison (rows 1-16) --------------
    section_title(ws, 1, "Extraction — contract comparison (p1 vs p2)")
    put(ws, 2, 1, "Metric", bold=True)
    put(ws, 2, 2, "p1", bold=True)
    put(ws, 2, 3, "p2", bold=True)
    style_header_row(ws, 3, row=2)

    def er_sumifs(sum_col: str, contract_cell: str) -> str:
        return (f"=SUMIFS({ER}!${sum_col}$2:${sum_col}$1000,"
                f"{ER}!$H$2:$H$1000,{contract_cell},"
                f"{ER}!$R$2:$R$1000,\"ok\")")

    rows_a = [
        (3,  "Units run (parents extracted)",
             f"=COUNTIFS({ER}!$H$2:$H$1000,B$2,{ER}!$R$2:$R$1000,\"ok\")", N),
        (4,  "Facts staged",        er_sumifs("M", "B$2"), N),
        (5,  "Quarantined",         er_sumifs("O", "B$2"), N),
        (6,  "Grounding flags",     er_sumifs("P", "B$2"), N),
        (7,  "Mentions staged",     er_sumifs("N", "B$2"), N),
        (8,  "Prompt tokens",       er_sumifs("J", "B$2"), N),
        (9,  "Output tokens",       er_sumifs("K", "B$2"), N),
        (10, "Wall ms (total)",     er_sumifs("L", "B$2"), N),
        (11, "Facts per unit",      "=IFERROR(B4/B3,0)", DEC),
        (12, "Quarantine rate",     "=IFERROR(B5/(B4+B5),0)", PCT),
        (13, "Grounding flag rate", "=IFERROR(B6/B4,0)", PCT),
        (14, "Avg wall ms per unit","=IFERROR(B10/B3,0)", N),
        (15, "Avg tokens per unit", "=IFERROR((B8+B9)/B3,0)", N),
        (16, "Repairs",             er_sumifs("Q", "B$2"), N),
    ]
    for r, label, formula_b, fmt in rows_a:
        put(ws, r, 1, label)
        put(ws, r, 2, formula_b, fmt=fmt)
        put(ws, r, 3, formula_b.replace("B$2", "C$2").replace("B4", "C4")
                               .replace("B5", "C5").replace("B6", "C6")
                               .replace("B3", "C3").replace("B10", "C10")
                               .replace("B8", "C8").replace("B9", "C9"), fmt=fmt)

    # --- Section B: grounding outcomes (rows 18-25) --------------------------
    section_title(ws, 18, "Grounding outcomes (staged facts)")
    put(ws, 19, 1, "Outcome", bold=True)
    put(ws, 19, 2, "Count", bold=True)
    style_header_row(ws, 2, row=19)
    PF = "Raw_PendingFacts"
    for i, outcome in enumerate(["pass", "span_missing", "components_missing",
                                 "construction"]):
        r = 20 + i
        put(ws, r, 1, outcome)
        put(ws, r, 2, f"=COUNTIFS({PF}!$E$2:$E$1000,$A{r})", fmt=N)
    put(ws, 24, 1, "Grounding pass rate")
    put(ws, 24, 2, "=IFERROR(B20/SUM(B20:B23),0)", fmt=PCT)
    put(ws, 25, 1, "Facts promoted to fact store")
    put(ws, 25, 2, f"=COUNTIFS({PF}!$M$2:$M$1000,\"promoted\")", fmt=N)

    # --- Section C: quarantine by reason x contract (rows 27-31) -------------
    section_title(ws, 27, "Quarantine by reason")
    put(ws, 28, 1, "Reason", bold=True)
    put(ws, 28, 2, "p1", bold=True)
    put(ws, 28, 3, "p2", bold=True)
    style_header_row(ws, 3, row=28)
    QR = "Raw_Quarantine"
    for i, reason in enumerate(["unbound_entity_type", "unbound_predicate",
                                "validation_failure"]):
        r = 29 + i
        put(ws, r, 1, reason)
        for col, hdr in ((2, "B$28"), (3, "C$28")):
            put(ws, r, col,
                f"=COUNTIFS({QR}!$C$2:$C$1000,$A{r},{QR}!$G$2:$G$1000,{hdr})",
                fmt=N)

    # --- Section D: resolution tier x decision (rows 33-39) ------------------
    section_title(ws, 33, "Resolution decisions — tier × decision")
    put(ws, 34, 1, "Tier", bold=True)
    for col, d in ((2, "resolved"), (3, "new_entity"), (4, "review")):
        put(ws, 34, col, d, bold=True)
    put(ws, 34, 5, "avg wall ms", bold=True)
    style_header_row(ws, 5, row=34)
    RD = "Raw_Resolution"
    for i, tier in enumerate(["t0", "t1", "t1b", "none"]):
        r = 35 + i
        put(ws, r, 1, tier)
        for col in (2, 3, 4):
            hdr = f"{get_column_letter(col)}$34"
            put(ws, r, col,
                f"=COUNTIFS({RD}!$D$2:$D$1000,$A{r},{RD}!$H$2:$H$1000,{hdr})",
                fmt=N)
        put(ws, r, 5,
            f"=IFERROR(AVERAGEIFS({RD}!$K$2:$K$1000,{RD}!$D$2:$D$1000,$A{r}),0)",
            fmt=N)

    # --- Section E: match candidate score distribution (rows 41-47) ----------
    section_title(ws, 41, "Match-candidate score distribution")
    put(ws, 42, 1, "Score bin", bold=True)
    put(ws, 42, 2, "Pairs", bold=True)
    style_header_row(ws, 2, row=42)
    MC = "Raw_MatchCandidates"
    bins = [("0.0 – 0.2", ">=0", "<0.2"), ("0.2 – 0.4", ">=0.2", "<0.4"),
            ("0.4 – 0.6", ">=0.4", "<0.6"), ("0.6 – 0.8", ">=0.6", "<0.8"),
            ("0.8 – 1.0", ">=0.8", None)]
    for i, (label, lo, hi) in enumerate(bins):
        r = 43 + i
        put(ws, r, 1, label)
        cond = f"{MC}!$G$2:$G$5000,\"{lo}\""
        if hi:
            cond += f",{MC}!$G$2:$G$5000,\"{hi}\""
        put(ws, r, 2, f"=COUNTIFS({cond})", fmt=N)

    # --- Section F: review queue by kind (rows 49-56) -------------------------
    section_title(ws, 49, "Review queue (open items by kind)")
    put(ws, 50, 1, "Kind", bold=True)
    put(ws, 50, 2, "Items", bold=True)
    style_header_row(ws, 2, row=50)
    RQ = "Raw_ReviewQueue"
    kinds = ["mention", "match", "oversized_fact", "document", "quarantine",
             "pending_fact"]
    for i, kind in enumerate(kinds):
        r = 51 + i
        put(ws, r, 1, kind)
        put(ws, r, 2, f"=COUNTIFS({RQ}!$A$2:$A$1000,$A{r})", fmt=N)
    put(ws, 57, 1, "Total open", bold=True)
    put(ws, 57, 2, "=SUM(B51:B56)", bold=True, fmt=N)

    # --- Section G: flywheel labels (rows 59-64) -------------------------------
    section_title(ws, 59, "Label flywheel (gold-set fuel accumulated)")
    put(ws, 60, 1, "label_type / source", bold=True)
    put(ws, 60, 2, "Count", bold=True)
    style_header_row(ws, 2, row=60)
    LB = "Raw_Labels"
    combos = [("er_match", "deterministic"), ("er_match", "human_review"),
              ("er_nonmatch", "human_review"), ("er_nonmatch", "reversal")]
    for i, (lt, src) in enumerate(combos):
        r = 61 + i
        put(ws, r, 1, f"{lt} / {src}")
        put(ws, r, 2,
            f"=COUNTIFS({LB}!$C$2:$C$1000,\"{lt}\",{LB}!$D$2:$D$1000,\"{src}\")",
            fmt=N)
    put(ws, 65, 1, "Total labels", bold=True)
    put(ws, 65, 2, f"=COUNTA({LB}!$A$2:$A$1000)", bold=True, fmt=N)

    return ws


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
def _bar(ws_dash, title: str, y_title: str, anchor: str, width=15.5, height=8.5):
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "clustered"
    ch.title = title
    ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.gapWidth = 80
    ch.width = width
    ch.height = height
    ws_dash.add_chart(ch, anchor)
    return ch


def _series(m, col: str, r1: int, r2: int, title: str) -> Series:
    ref = Reference(m, min_col={"B": 2, "C": 3, "D": 4}[col], min_row=r1, max_row=r2)
    return Series(ref, title=title)


def _color(series: Series, hexval: str):
    series.graphicalProperties = GraphicalProperties(solidFill=hexval)


def build_dashboard(wb: Workbook):
    ws = wb.create_sheet("Dashboard", 1)
    ws.sheet_view.showGridLines = False
    m = wb["Metrics"]
    sheet_title(ws, "Knowledge Hub — pipeline metrics dashboard",
                "All numbers recompute from the Raw_ sheets. Re-run "
                "metrics_report.py for a fresh snapshot.", span=14)

    # KPI tiles (rows 4-5) — values are formulas into Metrics. Each tile
    # spans 2 columns with a spacer column between tiles.
    kpis = [
        ("FACTS PROMOTED",        "=Metrics!B25",  "#,##0"),
        ("GROUNDING PASS RATE",   "=Metrics!B24",  "0.0%"),
        ("OPEN REVIEW ITEMS",     "=Metrics!B57",  "#,##0"),
        ("FLYWHEEL LABELS",       "=Metrics!B65",  "#,##0"),
        ("QUARANTINE RATE (P2)",  "=Metrics!C12",  "0.0%"),
    ]
    for i, (label, formula, fmt) in enumerate(kpis):
        draw_tile(ws, 4, 1 + i * 3, label, formula, fmt)
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 34

    for col_idx in range(1, 16):
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            5 if col_idx % 3 == 0 else 12)   # spacer columns between tiles

    # Chart 1 — contract p1 vs p2 outcomes (Metrics rows 4-6).
    ch = _bar(ws, "Extraction outcomes — contract p1 vs p2", "items", "A8")
    for col, name, color in (("B", "p1", SERIES_1), ("C", "p2", SERIES_2)):
        s = _series(m, col, 4, 6, name)
        _color(s, color)
        ch.series.append(s)
    ch.set_categories(Reference(m, min_col=1, min_row=4, max_row=6))

    # Chart 2 — avg wall ms per unit (row 14), categories = contracts (row 2).
    ch = _bar(ws, "Cost per parent unit — wall clock", "ms / unit", "J8")
    ch.add_data(Reference(m, min_col=2, max_col=3, min_row=14, max_row=14),
                from_rows=True, titles_from_data=False)
    ch.set_categories(Reference(m, min_col=2, max_col=3, min_row=2, max_row=2))
    _color(ch.series[0], SERIES_1)
    ch.legend = None

    # Chart 3 — avg tokens per unit (row 15).
    ch = _bar(ws, "Cost per parent unit — tokens", "tokens / unit", "J26")
    ch.add_data(Reference(m, min_col=2, max_col=3, min_row=15, max_row=15),
                from_rows=True, titles_from_data=False)
    ch.set_categories(Reference(m, min_col=2, max_col=3, min_row=2, max_row=2))
    _color(ch.series[0], SERIES_1)
    ch.legend = None

    # Chart 4 — grounding outcomes (rows 20-23).
    ch = _bar(ws, "Grounding outcomes (staged facts)", "facts", "A26")
    ch.add_data(Reference(m, min_col=2, min_row=20, max_row=23),
                titles_from_data=False)
    ch.set_categories(Reference(m, min_col=1, min_row=20, max_row=23))
    _color(ch.series[0], SERIES_1)
    ch.legend = None

    # Chart 5 — quarantine by reason, p1 vs p2 (rows 29-31).
    ch = _bar(ws, "Quarantine by reason — contract p1 vs p2", "items", "A44")
    for col, name, color in (("B", "p1", SERIES_1), ("C", "p2", SERIES_2)):
        s = _series(m, col, 29, 31, name)
        _color(s, color)
        ch.series.append(s)
    ch.set_categories(Reference(m, min_col=1, min_row=29, max_row=31))

    # Chart 6 — resolution tier x decision (rows 35-38, series = decisions).
    ch = _bar(ws, "Resolution decisions by tier", "mentions", "J44")
    for col, name, color in (("B", "resolved", SERIES_1),
                             ("C", "new_entity", SERIES_2),
                             ("D", "review", SERIES_3)):
        s = _series(m, col, 35, 38, name)
        _color(s, color)
        ch.series.append(s)
    ch.set_categories(Reference(m, min_col=1, min_row=35, max_row=38))

    # Chart 7 — match-candidate score distribution (rows 43-47).
    ch = _bar(ws, "Match-candidate score distribution", "pairs", "A62")
    ch.add_data(Reference(m, min_col=2, min_row=43, max_row=47),
                titles_from_data=False)
    ch.set_categories(Reference(m, min_col=1, min_row=43, max_row=47))
    _color(ch.series[0], SERIES_1)
    ch.legend = None

    return ws


def build_benchmarks_sheet(wb: Workbook, board_rows: list[tuple]):
    """Axis leaderboards. One chart per (axis, gold_set, pin_profile) group —
    the methodology's comparability rule made physical: configs are only ever
    charted against runs under identical conditions."""
    ws = wb.create_sheet("Benchmarks")
    ws.sheet_view.showGridLines = False
    lb = wb["Raw_Leaderboard"]
    sheet_title(ws, "Benchmark leaderboards",
                "One chart = one (axis, gold set, pin profile) — runs are "
                "never compared across conditions.", span=14)
    put(ws, 3, 1, "'(advisory)' = below the methodology's statistical floors "
                  "(or backfilled): recorded, but cannot decide a winner.",
        color=MUTED, size=9)
    for col_idx in range(1, 16):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    # Group contiguous leaderboard rows (the SQL orders by the group key).
    # Raw_Leaderboard columns: B=axis, C=gold_set, D=pin_profile, F=config,
    # G=headline_name, H=headline_value; data starts at sheet row 2.
    groups: list[tuple[tuple, int, int]] = []   # (key, first_row, last_row)
    for ix, row in enumerate(board_rows, start=2):
        key = (row[1], row[2], row[3])
        if groups and groups[-1][0] == key:
            groups[-1] = (key, groups[-1][1], ix)
        else:
            groups.append((key, ix, ix))

    anchor_row = 5
    for (axis, gold_set, pins), r1, r2 in groups:
        headline = board_rows[r1 - 2][6] or "headline"
        ch = _bar(ws, f"{axis} — {gold_set} @ {pins}", headline,
                  f"A{anchor_row}")
        ch.add_data(Reference(lb, min_col=8, min_row=r1, max_row=r2),
                    titles_from_data=False)
        ch.set_categories(Reference(lb, min_col=6, min_row=r1, max_row=r2))
        _color(ch.series[0], SERIES_1)
        ch.legend = None
        anchor_row += 18
    return ws


def build_readme(wb: Workbook, dsn_display: str, counts: dict[str, int],
                 has_benchmark: bool):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 88
    sheet_title(ws, "Knowledge Hub — metrics report", span=2)

    r = 3

    def banner(text: str):
        nonlocal r
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws.cell(row=r, column=1, value=text.upper())
        cell.font = Font(name=FONT, size=10, bold=True, color=INK_2)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = Border(bottom=HAIRLINE, top=HAIRLINE)
        cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 20
        r += 1

    def kv(key: str, value, *, muted=False):
        nonlocal r
        kc = ws.cell(row=r, column=1, value=key)
        kc.font = Font(name=FONT, size=10, bold=True, color=INK)
        kc.alignment = Alignment(vertical="top", indent=1)
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = Font(name=FONT, size=10, color=INK_2 if muted else INK)
        vc.alignment = Alignment(vertical="top", wrap_text=True)
        if isinstance(value, str) and len(value) > 90:
            ws.row_dimensions[r].height = 13 * (len(value) // 85 + 1)
        r += 1

    banner("Snapshot")
    kv("Generated at", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S (local)"))
    kv("Source", dsn_display)
    kv("Script", f"metrics_report.py v{SCRIPT_VERSION}")

    banner("How to read this workbook")
    kv("Dashboard", "KPI tiles + charts. Everything is a formula over Metrics, "
                    "which is formulas over the Raw_ sheets — nothing is a "
                    "pasted number.")
    if has_benchmark:
        kv("Benchmarks", "Axis leaderboards — one chart per (axis, gold set, "
                         "pin profile); runs are never compared across "
                         "conditions. '(advisory)' bars cannot decide winners.")
    kv("Metrics", "COUNTIFS/SUMIFS aggregations. Edit/extend Raw_ rows and "
                  "these recompute in Excel.")
    kv("Raw_* sheets", "Verbatim snapshot of the pilot DB observability tables "
                       "at generation time. Re-run the script for fresh data.")

    banner("Assumptions")
    kv("contract column", "Derived from extractor_version: values without a "
                          "'/pN' suffix predate the suffix convention and are "
                          "labeled 'p1' (EXTRACTION_NOTES.md documents the "
                          "p1->p2 story).")
    kv("p1 vs p2", "Same model weights (qwen3.6@07d35212591f), different "
                   "prompt contracts — the workbook's headline comparison. "
                   "p1's near-total quarantine is REAL history, not an error; "
                   "the machinery caught a contract bug.")
    kv("Extraction metrics", "Only status='ok' runs are counted.")

    banner("Row counts at generation")
    for name, n in counts.items():
        kv(name, n, muted=True)

    banner("Benchmark harness")
    if has_benchmark:
        kv("Status", "migration 006 applied — see the Benchmarks sheet and "
                     "Raw_BenchmarkRuns / Raw_Leaderboard. Advisory runs are "
                     "recorded but cannot decide winners.")
    else:
        kv("Status", "benchmark_runs not present yet (migration 006 pending). "
                     "When it lands, this pipeline adds the sheets "
                     "automatically.")


# ---------------------------------------------------------------------------
# Recalculation. openpyxl writes formulas with NO cached values, and
# LibreOffice Calc does not recalculate xlsx on open — so a freshly written
# report shows blanks in Calc until something computes it. We drive a
# headless LibreOffice (throwaway profile, own port, whole process tree
# killed afterwards so a user's open Calc windows are never touched).
# ---------------------------------------------------------------------------
SOFFICE = Path(r"C:\Program Files\LibreOffice\program\soffice.exe")


def recalc_with_libreoffice(path: Path) -> bool:
    import shutil

    lo_python = SOFFICE.parent / "python.exe"
    helper = Path(__file__).parent / "recalc_uno.py"
    if not (SOFFICE.exists() and lo_python.exists() and helper.exists()):
        return False
    # A crashed/killed soffice leaves a '.~lock.<file>#' beside the workbook;
    # LibreOffice then refuses to open it (loadComponentFromURL returns None)
    # and recalc silently no-ops. Clear our own report's stale lock first —
    # it can only be ours, since we never leave a Calc window open on it.
    lock = path.with_name(f".~lock.{path.name}#")
    if lock.exists():
        try:
            lock.unlink()
        except OSError:
            pass
    profile = Path(tempfile.mkdtemp(prefix="kh_lo_profile_"))
    proc = subprocess.Popen(
        [str(SOFFICE), "--headless", "--norestore", "--nologo",
         f"-env:UserInstallation={profile.as_uri()}",
         "--accept=socket,host=localhost,port=2002;urp;"])
    try:
        result = subprocess.run([str(lo_python), str(helper), str(path)],
                                capture_output=True, text=True, timeout=180)
        ok = result.returncode == 0 and "RECALC OK" in result.stdout
        if not ok:
            print(f"recalc failed: {result.stdout} {result.stderr}".strip())
        return ok
    finally:
        try:  # kill the whole tree: soffice.exe launcher + soffice.bin child
            import psutil
            root = psutil.Process(proc.pid)
            for child in root.children(recursive=True):
                child.kill()
            root.kill()
        except Exception:
            proc.kill()
        shutil.rmtree(profile, ignore_errors=True)   # don't accrete temp profiles


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=None, help="output .xlsx path")
    ap.add_argument("--dsn", default=None, help="Postgres DSN override")
    ap.add_argument("--tenant", default="default",
                    help="tenant scope for the pipeline-observability sheets "
                         "(benchmark sheets stay cross-tenant; they carry "
                         "tenant labels)")
    ap.add_argument("--no-recalc", action="store_true",
                    help="skip the LibreOffice recalculation pass")
    args = ap.parse_args()

    dsn = args.dsn or _dsn()
    out = Path(args.out) if args.out else (
        Path(__file__).parent / "reports" /
        f"KnowledgeHub_Metrics_{dt.date.today().isoformat()}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we create our own

    counts: dict[str, int] = {}
    with psycopg.connect(dsn, connect_timeout=10) as conn:
        has_benchmark = conn.execute(
            "SELECT EXISTS (SELECT 1 FROM information_schema.tables "
            "WHERE table_name = 'benchmark_runs')").fetchone()[0]
        sheets = dict(RAW_SHEETS)
        if has_benchmark:
            sheets.update(BENCHMARK_SHEETS)
        board_rows: list[tuple] = []
        for name, (headers, sql) in sheets.items():
            params = {"tenant": args.tenant} if "%(tenant)s" in sql else None
            rows = conn.execute(sql, params).fetchall()
            counts[name] = write_raw_sheet(wb, name, headers, rows)
            if name == "Raw_Leaderboard":
                board_rows = rows

    build_metrics_sheet(wb)
    build_dashboard(wb)
    if board_rows:
        build_benchmarks_sheet(wb, board_rows)
    # Mask credentials in the display DSN — both URL form
    # (postgresql://user:PASSWORD@host/db) and key=value form (password=...).
    import re
    dsn_display = re.sub(r"(://[^:/@]+):[^@]+@", r"\1:***@", dsn)
    dsn_display = " ".join(p for p in dsn_display.split()
                           if not p.startswith("password="))
    build_readme(wb, dsn_display, counts, has_benchmark)

    # Sheet order: README, Dashboard, Benchmarks (when present), Metrics, Raw_*
    order = ["README", "Dashboard"] + (
        ["Benchmarks"] if "Benchmarks" in wb.sheetnames else []) + ["Metrics"] + [
        s for s in wb.sheetnames if s.startswith("Raw_")]
    wb._sheets = [wb[s] for s in order]

    wb.save(out)
    print(f"report written: {out}")
    for name, n in counts.items():
        print(f"  {name}: {n} rows")

    if not args.no_recalc:
        if recalc_with_libreoffice(out):
            print("formulas recalculated (LibreOffice headless)")
        else:
            print("WARNING: recalculation skipped — open the file in Excel "
                  "(recalcs on open) or press Ctrl+Shift+F9 in Calc")
    return 0


if __name__ == "__main__":
    sys.exit(main())
